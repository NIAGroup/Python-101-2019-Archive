import os
#import openpyxl
from openpyxl.styles import PatternFill, Font

currentDir = os.getcwd()
currentDir = currentDir + "\\ParserX_Dir"
print("Current search directory is: " + currentDir)
#targetFile = currentDir + "\\Intel_Core_Processors1.xlsx"
#print(targetFile)

from os import listdir
files = listdir(currentDir) # files is a list
number_files = len(files)

print("There are %s" % number_files + " files in the directory.")

excel_docs = []      # .xlsx docs to consolodate into master file

for i in range(number_files):
    current_file = files[i]
    delimiter_position = current_file.find('.')
    if ( (current_file[delimiter_position:] == ".xlsx") and not("Master" in current_file) ):
        excel_docs.append(current_file)

print(excel_docs)        

# search each file for lines containing 5th
# write each in into the Master file
# Then do likewise for 6th, 7th, 8th, and 9th

generation_numbers = ["5th", "6th", "7th", "8th", "9th"]
from  openpyxl import load_workbook

# create the Master consolodation worksheet
from openpyxl import Workbook
MasterFileName = currentDir + "\\Master_Intel_Processors.xlsx"
newWorkBook = Workbook()
ws1 = newWorkBook.active
ws1.title = "Master List"

MasterFile_RowCount = 0
ItemFoundRowCount = 0

num_row_string_found_in_source = 0
        
for k_gen_number in range(len(generation_numbers)):
    print(" ")
    for i_excel_doc in range(len(excel_docs)): # len(excel_docs)
        wb = load_workbook(currentDir + "\\" + excel_docs[i_excel_doc], "r")
        sheet = wb.worksheets[0]
        row_count = sheet.max_row
        #ItemFoundRowCount = ItemFoundRowCount + row_count
        column_count = sheet.max_column
        #MasterFile_RowCount = MasterFile_RowCount + 1

        if(i_excel_doc == 0):
            print("Search string is \"" + generation_numbers[k_gen_number] + "\"")
            
        # Print out how many rows are in each spreadsheet
        print(excel_docs[i_excel_doc] + " has " + str(row_count) + " rows. Generation list index is " + str(k_gen_number))
        sheet.cell(1, 1)
        
        # iterate through the rows in the spreadsheet
        row_noString = 0

        for j_row in range(1, row_count + 1):
            cell_obj = sheet.cell(row =j_row, column = 7)

            if k_gen_number == 0 and j_row == 1 and i_excel_doc == 0:
                print("First row in first file.")
                for j1_col in range(1, column_count + 1):
                    cell_obj2 = sheet.cell(row =j_row, column = j1_col)
                    #print(cell_obj2.value)
                    something2 = ws1.cell(row = j_row, column = j1_col, value= cell_obj2.value)
                    something2.font = Font(bold=True)
                num_row_string_found_in_source = num_row_string_found_in_source + 1    
                
            if generation_numbers[k_gen_number] in str(cell_obj.value) :
                num_row_string_found_in_source = num_row_string_found_in_source + 1                
                #print("Target Row number = " + str(num_row_string_found_in_source) + " with value = " + str(cell_obj.value))
                something = ws1.cell(row = num_row_string_found_in_source, column = 1, value= cell_obj.value)
                for j_col in range(1, column_count + 1):
                    cell_obj = sheet.cell(row = j_row, column = j_col)
                    if j_col == 1:
                        ws1.column_dimensions["A"].width = 55
                    if j_col == 2:
                        ws1.column_dimensions["B"].width = 25
                    if j_col == 3:
                        ws1.column_dimensions["C"].width = 10
                    if j_col == 4:
                        ws1.column_dimensions["D"].width = 15
                    if j_col == 5:
                        ws1.column_dimensions["E"].width = 35
                    if j_col == 6:
                        ws1.column_dimensions["F"].width = 25
                    if j_col == 7:
                        ws1.column_dimensions["G"].width = 20
                    copy_to_target_cell = ws1.cell(row = num_row_string_found_in_source, column = j_col, value= cell_obj.value)
                    if j_col == column_count:
                        if k_gen_number == 0:
                            copy_to_target_cell.fill = PatternFill("solid", fgColor="00EEEE")
                        if k_gen_number == 1:
                            copy_to_target_cell.fill = PatternFill("solid", fgColor="FF83FA")
                        if k_gen_number == 2:
                            copy_to_target_cell.fill = PatternFill("solid", fgColor="0000FF")
                        if k_gen_number == 3:
                            copy_to_target_cell.fill = PatternFill("solid", fgColor="9AFF9A")                            
                        if k_gen_number == 4:
                            copy_to_target_cell.fill = PatternFill("solid", fgColor="FF0000")
# save the newly created master Excel WorkBook
newWorkBook.save(MasterFileName)    





